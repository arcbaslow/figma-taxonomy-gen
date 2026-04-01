"""Tests for output formatters: Excel, CSV, JSON Schema, Markdown."""

import json
from pathlib import Path

import pytest

from figma_taxonomy.config import load_config
from figma_taxonomy.extractor import extract_elements
from figma_taxonomy.models import TaxonomyEvent, EventProperty
from figma_taxonomy.taxonomy_engine import generate_taxonomy


FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture
def config():
    return load_config(None)


@pytest.fixture
def taxonomy():
    with open(FIXTURES_DIR / "banking_app.json") as f:
        figma_file = json.load(f)
    config = load_config(None)
    elements = extract_elements(figma_file, config)
    return generate_taxonomy(elements, config)


@pytest.fixture
def sample_events():
    """Minimal hand-crafted events for output testing."""
    return [
        TaxonomyEvent(
            event_name="login_pageview",
            flow="Authentication",
            description="User views login screen",
            properties=[
                EventProperty(name="screen_name", type="string", description="Screen where event occurred"),
            ],
            source_node_id="",
        ),
        TaxonomyEvent(
            event_name="login_submit_clicked",
            flow="Authentication",
            description="User clicks submit button",
            properties=[
                EventProperty(name="element_text", type="string", description="Button text"),
                EventProperty(name="screen_name", type="string", description="Screen where event occurred"),
            ],
            source_node_id="1:14",
        ),
        TaxonomyEvent(
            event_name="payment_amount_entered",
            flow="Payments",
            description="User enters payment amount",
            properties=[
                EventProperty(name="field_name", type="string", description="Field name"),
                EventProperty(name="is_valid", type="boolean", description="Validation passed"),
                EventProperty(name="screen_name", type="string", description="Screen"),
            ],
            source_node_id="3:10",
        ),
    ]


# --- Excel tests ---

class TestExcelOutput:
    def test_creates_xlsx_file(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.excel import write_excel

        output_path = tmp_path / "taxonomy.xlsx"
        write_excel(sample_events, config, output_path)
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_has_events_sheet(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.excel import write_excel
        import openpyxl

        output_path = tmp_path / "taxonomy.xlsx"
        write_excel(sample_events, config, output_path)

        wb = openpyxl.load_workbook(output_path)
        assert "Events" in wb.sheetnames

    def test_has_parameters_sheet(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.excel import write_excel
        import openpyxl

        output_path = tmp_path / "taxonomy.xlsx"
        write_excel(sample_events, config, output_path)

        wb = openpyxl.load_workbook(output_path)
        assert "Parameters" in wb.sheetnames

    def test_events_sheet_headers(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.excel import write_excel
        import openpyxl

        output_path = tmp_path / "taxonomy.xlsx"
        write_excel(sample_events, config, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Events"]
        headers = [ws.cell(row=2, column=c).value for c in range(2, 14)]
        assert headers[0] == "Flow"
        assert headers[1] == "Event Name"
        assert headers[2] == "Event Description"
        assert headers[3] == "Parameter Set"

    def test_events_sheet_data_rows(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.excel import write_excel
        import openpyxl

        output_path = tmp_path / "taxonomy.xlsx"
        write_excel(sample_events, config, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Events"]

        # Row 3 is first data row (row 1 empty, row 2 headers)
        assert ws.cell(row=3, column=2).value == "Authentication"
        assert ws.cell(row=3, column=3).value == "login_pageview"
        assert ws.cell(row=3, column=4).value == "User views login screen"

    def test_parameters_sheet_has_global_params(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.excel import write_excel
        import openpyxl

        output_path = tmp_path / "taxonomy.xlsx"
        write_excel(sample_events, config, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Parameters"]
        param_names = [ws.cell(row=r, column=2).value for r in range(3, 10)]
        assert "screen_name" in param_names
        assert "platform" in param_names

    def test_full_taxonomy_output(self, taxonomy, config, tmp_path):
        from figma_taxonomy.output.excel import write_excel
        import openpyxl

        output_path = tmp_path / "taxonomy.xlsx"
        write_excel(taxonomy, config, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Events"]
        # Should have header row + data rows
        data_rows = [
            ws.cell(row=r, column=3).value
            for r in range(3, ws.max_row + 1)
            if ws.cell(row=r, column=3).value
        ]
        assert len(data_rows) == len(taxonomy)


# --- CSV tests ---

class TestAmplitudeCsvOutput:
    def test_creates_csv_file(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.amplitude_csv import write_csv

        output_path = tmp_path / "taxonomy.csv"
        write_csv(sample_events, config, output_path)
        assert output_path.exists()

    def test_csv_headers(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.amplitude_csv import write_csv

        output_path = tmp_path / "taxonomy.csv"
        write_csv(sample_events, config, output_path)

        lines = output_path.read_text().strip().split("\n")
        assert lines[0] == "Event Type,Category,Description,Property Name,Property Type,Property Description"

    def test_csv_data_rows(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.amplitude_csv import write_csv

        output_path = tmp_path / "taxonomy.csv"
        write_csv(sample_events, config, output_path)

        lines = output_path.read_text().strip().split("\n")
        # Each event-property combination is a row
        assert len(lines) > 1  # header + data


# --- JSON tests ---

class TestJsonSchemaOutput:
    def test_creates_json_file(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.json_schema import write_json

        output_path = tmp_path / "taxonomy.json"
        write_json(sample_events, config, output_path)
        assert output_path.exists()

    def test_json_structure(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.json_schema import write_json

        output_path = tmp_path / "taxonomy.json"
        write_json(sample_events, config, output_path)

        data = json.loads(output_path.read_text())
        assert "$schema" in data
        assert "events" in data
        assert "login_pageview" in data["events"]
        assert "login_submit_clicked" in data["events"]

    def test_json_event_properties(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.json_schema import write_json

        output_path = tmp_path / "taxonomy.json"
        write_json(sample_events, config, output_path)

        data = json.loads(output_path.read_text())
        event = data["events"]["login_submit_clicked"]
        assert event["category"] == "Authentication"
        assert "properties" in event
        assert "element_text" in event["properties"]


# --- Markdown tests ---

class TestMarkdownOutput:
    def test_creates_md_file(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.markdown import write_markdown

        output_path = tmp_path / "taxonomy.md"
        write_markdown(sample_events, config, output_path)
        assert output_path.exists()

    def test_markdown_has_flow_headers(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.markdown import write_markdown

        output_path = tmp_path / "taxonomy.md"
        write_markdown(sample_events, config, output_path)

        content = output_path.read_text()
        assert "## Authentication" in content
        assert "## Payments" in content

    def test_markdown_has_event_sections(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.markdown import write_markdown

        output_path = tmp_path / "taxonomy.md"
        write_markdown(sample_events, config, output_path)

        content = output_path.read_text()
        assert "### login_pageview" in content
        assert "### login_submit_clicked" in content

    def test_markdown_lists_properties(self, sample_events, config, tmp_path):
        from figma_taxonomy.output.markdown import write_markdown

        output_path = tmp_path / "taxonomy.md"
        write_markdown(sample_events, config, output_path)

        content = output_path.read_text()
        assert "`element_text`" in content
        assert "`field_name`" in content
