"""Excel output formatter matching the taxonomy template structure."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from figma_taxonomy.config import TaxonomyConfig
from figma_taxonomy.models import TaxonomyEvent


def write_excel(
    events: list[TaxonomyEvent],
    config: TaxonomyConfig,
    output_path: Path,
) -> None:
    wb = openpyxl.Workbook()

    # --- Events sheet ---
    ws_events = wb.active
    ws_events.title = "Events"

    headers = [
        "", "Flow", "Event Name", "Event Description", "Parameter Set",
        "Parameter Name", "Parameter Description",
        "Parameter Name", "Parameter Description",
        "Parameter Name", "Parameter Description",
        "Parameter Name", "Parameter Description",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_events.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    global_names = {p["name"] for p in config.global_properties}

    for i, event in enumerate(events):
        row = i + 3
        ws_events.cell(row=row, column=2, value=event.flow)
        ws_events.cell(row=row, column=3, value=event.event_name)
        ws_events.cell(row=row, column=4, value=event.description)

        event_props = [p for p in event.properties if p.name not in global_names]

        if event_props:
            param_set = ", ".join(p.name for p in event_props)
            ws_events.cell(row=row, column=5, value=param_set)

            for j, prop in enumerate(event_props[:4]):
                name_col = 6 + (j * 2)
                desc_col = 7 + (j * 2)
                ws_events.cell(row=row, column=name_col, value=prop.name)
                ws_events.cell(row=row, column=desc_col, value=prop.description)

    # --- Parameters sheet ---
    ws_params = wb.create_sheet("Parameters")

    param_headers = [
        "", "Parameter Name", "Parameter Description",
        "Parameter Name", "Parameter Description",
        "Parameter Name", "Parameter Description",
        "Parameter Name", "Parameter Description",
    ]
    for col, header in enumerate(param_headers, 1):
        cell = ws_params.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    for i, prop in enumerate(config.global_properties):
        row = i + 3
        ws_params.cell(row=row, column=2, value=prop["name"])
        ws_params.cell(row=row, column=3, value=prop.get("description", ""))

    wb.save(output_path)
